import os

import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import shutil
from tensorboard.backend.event_processing.event_accumulator import EventAccumulator


def create_spreadsheet(target_epochs =-1, template_path="templates/ResultsExcelTemplate.xlsx"):
    """ Creates a spreadsheet with the best and the end result from each run with the given target number
        of epochs if one is given for all runs in the runs folder.
        Automatically adds an image with a plot of all the runs with markings for the best values.
        """
    # Load spreadsheet to enter data into
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    # Collect data, prepare plot
    run_names, best_values, end_values = _iterate_runs(target_epochs=target_epochs, values_list_handler=create_run_chart)

    # Write data into spreadsheet
    cnt = 0
    for run_names, best_values, end_values in zip(run_names, best_values, end_values):
        row_number = str(cnt + 7)

        cell_name_run_name = 'E' + row_number
        ws[cell_name_run_name] = run_names

        cell_name_end_value = 'F' + row_number
        ws[cell_name_end_value] = end_values

        cell_name_best_value = 'G' + row_number
        ws[cell_name_best_value] = best_values

        cnt += 1

    # Get the name of the last run to use it as a suffix for the results file and the plot image
    path = "runs"
    dir_list = os.listdir(path)
    suffix = dir_list[0]
    suffix = suffix.split("_", 3)
    suffix = suffix[3].rsplit("_", 2)[0]

    # Add chart to spreadsheet
    plot_name = 'plot_'+suffix+'.png'
    plt.savefig(plot_name)
    plt.close()

    # Save spreadsheet
    img = openpyxl.drawing.image.Image(plot_name)
    img.anchor = 'E92'
    ws.add_image(img)

    wb.save("results_" + suffix + ".xlsx")


def evaluate_runs(target_epochs =-1):
    """ Determines the best and the end result from each run with the given target number of epochs if one if given.
        In the end mean, sd, max, min and max-min of all best values are given.
        Skips all runs with deviating number of epochs and crashed runs giving an error message.
        If no target number of epochs or a negative value is given each run that is not giving an index error is evaluated.
     """
    def print_header_line():
        print("Name of Measurement                End Value         Best Value")

    def print_run_name_and_values(run_name, values_str):
        print(run_name, values_str)

    def print_run_name_and_values_with_epoch_warning(run_name, values_str, num_epochs_of_run, target_epochs, run_path):
        print(run_name, values_str, "WARNING: run has", num_epochs_of_run, "Epochs but should have", target_epochs)

    def print_run_name_and_index_error_warning(run_name, run_path):
        print(run_name + " produces an index error.")

    def print_runs_statistics(best_values_list, values_list):
        best_values_array =  np.asarray(best_values_list)
        print("Based only on runs with right number of epochs:")
        print("Mean:   ", best_values_array.mean())
        # Using ddof=1 to estimate the standard deviation based on a sample like Excels STDEV function.
        print("SD:     ", best_values_array.std(ddof=1))
        print("Max:    ", best_values_array.max())
        print("Min:    ", best_values_array.min())
        print("Max-Min:", best_values_array.max() - best_values_array.min())

        # Find first occurrence of best value in each run and print highest.
        # This number gives an idea of how many epochs are necessary.
        highest_epoch_of_first_occurence_best_value = 0
        for i in range(len(values_list)):
            run_array = np.asarray(values_list[i])
            run_array = np.squeeze(run_array)
            best_value_of_run = best_values_list[i]
            epoch_of_best_value = np.where(run_array == best_value_of_run)[0][0]
            if epoch_of_best_value > highest_epoch_of_first_occurence_best_value:
                highest_epoch_of_first_occurence_best_value = epoch_of_best_value

        print("Highest epoch of best value:", highest_epoch_of_first_occurence_best_value)

    run_names, best_values, end_values = _iterate_runs(target_epochs=target_epochs,
                                                       preparation=print_header_line,
                                                       right_number_of_epochs_run_handler=print_run_name_and_values,
                                                       wrong_number_of_epochs_run_handler=print_run_name_and_values_with_epoch_warning,
                                                       error_handler=print_run_name_and_index_error_warning,
                                                       values_list_handler=print_runs_statistics
                                                       )
    return run_names, best_values, end_values


def remove_unfinished_runs(target_epochs =-1):
    """ First removes all crashed runs that lead to an index error.
    Then removes all runs that have a different ammount of epochs then the number in target_epochs
    if a number greater 0 is given."""
    def print_right_number_epochs(run_name, values_str):
        print(run_name, "has right ammount of epochs.")

    def delete_run_wrong_number_epochs(run_name, values_str, num_epochs_of_run, target_epochs, run_path):
        shutil.rmtree(run_path)
        print(run_name, "has", num_epochs_of_run, "epochs but should have", target_epochs, "and was therefore deleted")

    def delete_run_index_error(run_name, run_path):
        shutil.rmtree(run_path)
        print("Run", run_name, "produces an index error and was therefore deleted.")

    _iterate_runs(target_epochs=target_epochs,
                  right_number_of_epochs_run_handler=print_right_number_epochs,
                  wrong_number_of_epochs_run_handler=delete_run_wrong_number_epochs,
                  error_handler=delete_run_index_error,
                  )

def create_run_chart(best_values_list, values_list):
    plt.figure(figsize=(32, 12), dpi=80)
    plt.yticks(np.arange(0, 1, 0.05))
    plt.ylim([0, 1])
    plt.tick_params(labeltop=True, labelright=True)

    for i in range(len(values_list)):
        run_array = np.asarray(values_list[i])
        run_array = np.squeeze(run_array)
        plt.plot(run_array)

        best_value_of_run = best_values_list[i]
        # Get the first occurrence of the best value of the run in the run and mark with a cross.
        epoch_of_best_value = np.where(run_array == best_value_of_run)[0][0]
        plt.scatter(epoch_of_best_value, best_value_of_run, marker='+', s=4000, color='black')

    plt.ylabel('Accuracy')
    plt.xlabel('Epoch')

    plt.title('Accuracy vs. Epoch')
    plt.grid(True)

def show_runs_chart(target_epochs =-1):
    """ Creates a chart of the accuracy vs. the epochs for all runs in the runs directory."""
    _iterate_runs(target_epochs=target_epochs, values_list_handler=create_run_chart)

    plt.show()


def _get_test_accuracies_of_run(event_file_path):
    # Iterate over event files in folder that holds run data (should only be one file) and save each accuraciy of run in list.
    # The index is the epoch (with 0 being the first epoch)
    test_accuracies = []
    for event_file_name in os.listdir(event_file_path):
        if event_file_name == 'run.log':
            continue

        event_file_path = os.path.join(event_file_path, event_file_name)
        event_file_interator = EventAccumulator(event_file_path).Reload()

        # Get each test accuracy data point from the event file and save only the test value from the point in a list.
        for test_accuracy_data_point in event_file_interator.Scalars('data/test_acc'):
            test_accuracies.append(test_accuracy_data_point.value)

    test_accuracies_np_array = np.asarray(test_accuracies)

    return test_accuracies_np_array


def _default_handler(*args):
    pass


def _iterate_runs(target_epochs                      = -1,
                  preparation                        = _default_handler,
                  right_number_of_epochs_run_handler = _default_handler,
                  wrong_number_of_epochs_run_handler = _default_handler,
                  error_handler                      = _default_handler,
                  values_list_handler                = _default_handler
                  ):
    path = "runs"

    # Call parameter prepare funktion
    preparation()

    dir_list = os.listdir(path)
    dir_list.sort()

    run_names = []
    best_values = []
    end_values = []
    values_list = []

    for dir in dir_list:
        exp_path = os.path.join(path, dir)
        if(os.path.isdir(exp_path)):
            try:
                test_accuracies = _get_test_accuracies_of_run(exp_path)
                best_value = test_accuracies.max()
                end_value = test_accuracies[len(test_accuracies) - 1]
                values_str = str(end_value) + " " + str(best_value)
                # prepare values for evaluation in german Excel
                values_str = values_str.replace(".", ",")

                # if the number of accuracies is right or the check is deactivated call right_number_of_epochs_run_handler
                if len(test_accuracies) == target_epochs or target_epochs < 0:
                    right_number_of_epochs_run_handler(dir, values_str)
                    run_names.append(dir)
                    best_values.append(test_accuracies.max())
                    end_values.append(end_value)
                    values_list.append(test_accuracies)
                #else call wrong_number_of_epochs_run_handler
                else:
                    wrong_number_of_epochs_run_handler(dir, values_str, str(len(test_accuracies)), target_epochs, exp_path)
            # if an index error occurred, as is common when a run was suddenly interrupted in the beginning, call index_error_run_handler
            except IndexError as e:
                error_handler(dir, exp_path)
            except KeyError as e:
                error_handler(dir, exp_path)

    # If a run with enough epochs was found, call values_list_handler
    if len(best_values) > 0:
        print(len(best_values), "runs with the right number of epochs.")
        values_list_handler(best_values, values_list)

    return run_names, best_values, end_values

# Evaluates all runs that are not crashed, using no fixed number of epochs
if __name__ == '__main__':
    evaluate_runs()
    #create_chart()